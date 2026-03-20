"""
Create a simplified Excel file with only the essential VBA macros
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

def create_simplified_excel():
    """Create Excel file with simplified VBA macro code"""
    
    # Read the simplified VBA macro code
    macro_file = Path(__file__).parent / "Excel_Threshold_Explorer_Simplified.bas"
    
    if not macro_file.exists():
        print(f"Error: Simplified macro file not found at {macro_file}")
        return
    
    with open(macro_file, 'r', encoding='utf-8') as f:
        vba_code = f.read()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Simplified VBA Macro"
    
    # Styles
    title_font = Font(size=16, bold=True, color="FFFFFF")
    header_font = Font(size=12, bold=True)
    code_font = Font(name="Consolas", size=10)
    
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    code_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Excel Threshold Explorer - Simplified VBA Macro"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws.merge_cells('A1:D1')
    
    # Instructions
    instructions = [
        ("This simplified version contains only the essential macros:", header_font, header_fill),
        ("", None, None),
        ("• SetupThresholdExplorer - Sets up everything automatically", None, None),
        ("• UpdateAnalysis - Updates when you change the threshold", None, None),
        ("• HighlightMatrix - Applies color coding to the similarity matrix", None, None),
        ("", None, None),
        ("Quick Setup:", header_font, header_fill),
        ("1. Copy the code from column B below", None, None),
        ("2. Press Alt+F11 to open VBA Editor", None, None),
        ("3. Insert > Module", None, None),
        ("4. Paste the code", None, None),
        ("5. Press Alt+F8 and run SetupThresholdExplorer", None, None),
        ("", None, None),
        ("VBA Code (copy everything below):", header_font, header_fill)
    ]
    
    row = 3
    for text, font, fill in instructions:
        if text:
            ws[f'A{row}'] = text
            ws[f'A{row}'].font = font or Font(size=11)
            ws[f'A{row}'].fill = fill or PatternFill()
            ws[f'A{row}'].alignment = Alignment(vertical='top')
            ws[f'A{row}'].border = thin_border
            ws.merge_cells(f'A{row}:D{row}')
        row += 1
    
    # Add VBA code
    lines = vba_code.split('\n')
    
    for line in lines:
        if len(line) > 32767:
            chunks = [line[i:i+3000] for i in range(0, len(line), 3000)]
            for chunk in chunks:
                ws[f'B{row}'] = chunk
                ws[f'B{row}'].font = code_font
                ws[f'B{row}'].fill = code_fill
                ws[f'B{row}'].alignment = Alignment(vertical='top', wrap_text=True)
                ws[f'B{row}'].border = thin_border
                row += 1
        else:
            ws[f'B{row}'] = line
            ws[f'B{row}'].font = code_font
            ws[f'B{row}'].fill = code_fill
            ws[f'B{row}'].alignment = Alignment(vertical='top', wrap_text=True)
            ws[f'B{row}'].border = thin_border
            row += 1
    
    # Add quick reference
    row += 2
    ws[f'A{row}'] = "Quick Reference:"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].border = thin_border
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    quick_refs = [
        ("SetupThresholdExplorer", "Run this once to set up everything"),
        ("UpdateAnalysis", "Run after changing threshold value"),
        ("Alt+F11", "Open VBA Editor"),
        ("Alt+F8", "Open Macro Dialog"),
        ("F5", "Run macro in VBA Editor")
    ]
    
    for macro, desc in quick_refs:
        ws[f'A{row}'] = macro
        ws[f'B{row}'] = desc
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    
    # Freeze panes
    ws.freeze_panes = 'A20'
    
    # Save
    output_file = Path(__file__).parent / "Threshold_Explorer_Simplified.xlsx"
    wb.save(output_file)
    
    print(f"Created simplified Excel file: {output_file}")
    print("\nThis file contains only the essential macros:")
    print("- SetupThresholdExplorer (main setup)")
    print("- UpdateAnalysis (update when threshold changes)")
    print("- HighlightMatrix (color coding)")
    print("\nMuch simpler and easier to understand!")

if __name__ == "__main__":
    create_simplified_excel()
