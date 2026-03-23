import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation


def build_threshold_explorer_workbook(
    summary_df: pd.DataFrame,
    groups_df: pd.DataFrame,
    similarity_matrix: np.ndarray = None,
    product_names: list = None,
    max_heatmap_products: int = 500
) -> bytes:
    """
    Build a self-contained Excel workbook for threshold exploration.
    
    Uses native Excel features (Data Validation dropdown, VLOOKUP formulas, charts)
    with pre-computed group data at multiple thresholds. No VBA, no Power Pivot.
    
    Args:
        summary_df: DataFrame with columns [Threshold, Groups Found, Products in Groups, 
                    Largest Group Size, Avg Group Similarity, Singletons]
        groups_df: DataFrame with columns [Threshold, Group ID, Group Summary, Product Name, 
                   Group Size, Group Avg Similarity, ...additional columns]
        similarity_matrix: Optional NxN similarity matrix for heatmap visualization
        product_names: Optional list of product names (required if similarity_matrix provided)
        max_heatmap_products: Maximum products to include in heatmap (for performance)
    
    Returns:
        Excel workbook as bytes
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. Create Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    _write_summary_sheet(ws_summary, summary_df, header_fill, header_font, border_thin)
    
    # 2. Create Groups sheet
    ws_groups = wb.create_sheet("Groups", 1)
    _write_groups_sheet(ws_groups, groups_df, header_fill, header_font, border_thin)
    
    # 3. Create Dashboard sheet (user-facing)
    ws_dashboard = wb.create_sheet("Dashboard", 0)  # Insert at beginning
    _write_dashboard_sheet(ws_dashboard, summary_df, title_font, header_fill, header_font, border_thin)
    
    # 4. Optional: Create Similarity Heatmap sheet
    if similarity_matrix is not None and product_names is not None:
        n_products = len(product_names)
        if n_products <= max_heatmap_products:
            ws_heatmap = wb.create_sheet("Similarity Heatmap", 3)
            _write_heatmap_sheet(ws_heatmap, similarity_matrix, product_names, header_fill, header_font)
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _write_summary_sheet(ws, summary_df, header_fill, header_font, border_thin):
    """Write the Summary sheet with threshold statistics."""
    # Write title
    ws['A1'] = "Threshold Analysis Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # Write data starting at row 3
    for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Header row formatting
            if r_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            cell.border = border_thin
    
    # Auto-size columns
    for column_cells in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column_cells:
            # Skip merged cells
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        if column_letter:
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width


def _write_groups_sheet(ws, groups_df, header_fill, header_font, border_thin):
    """Write the Groups sheet with dynamic threshold filtering and evolution tracking."""
    # Check if this is evolution data (has 'In Group' column)
    is_evolution_data = 'In Group' in groups_df.columns
    
    # Write title
    title = "Group Evolution - Track Membership Across Thresholds" if is_evolution_data else "Group Membership - Filtered by Threshold"
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    # Instructions
    if is_evolution_data:
        ws['A2'] = "See how individual groups evolve as threshold changes - members drop out but groups persist"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:H2')
    else:
        ws['A2'] = "Groups automatically filter based on threshold selected in Dashboard sheet"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:H2')
    
    # Threshold display (linked to Dashboard)
    ws['A3'] = "Current Threshold:"
    ws['A3'].font = Font(bold=True, size=11)
    ws['B3'] = "=Dashboard!$B$4"
    ws['B3'].font = Font(bold=True, size=11, color="4472C4")
    ws['B3'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws['B3'].border = border_thin
    ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Write all data starting at row 5 (for filtering)
    for r_idx, row in enumerate(dataframe_to_rows(groups_df, index=False, header=True), start=5):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Header row formatting
            if r_idx == 5:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            cell.border = border_thin
    
    # Add dynamic filtering using Excel formulas
    last_row = len(groups_df) + 5
    
    if is_evolution_data:
        # For evolution data, filter by both threshold and membership status
        ws['I5'] = "Filter"
        ws['I5'].fill = header_fill
        ws['I5'].font = header_font
        ws['I5'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add filter formula to all data rows
        for row_idx in range(6, last_row + 1):
            # Check if threshold matches AND product is in group
            # Column H contains 'In Group' values (TRUE/FALSE)
            ws[f'I{row_idx}'] = f'=IF(AND(C{row_idx}=Dashboard!$B$4, H{row_idx}=TRUE), 1, 0)'
        
        # Add conditional formatting for membership status
        for row_idx in range(6, last_row + 1):
            # Active members (green)
            ws.conditional_formatting.add(
                f'A{row_idx}:H{row_idx}',
                FormulaRule(
                    formula=[f'=$I{row_idx}=1'],
                    stopIfTrue=True,
                    fill=PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                )
            )
            
            # Inactive members at this threshold (light gray)
            ws.conditional_formatting.add(
                f'A{row_idx}:H{row_idx}',
                FormulaRule(
                    formula=[f'AND(C{row_idx}=Dashboard!$B$4, I{row_idx}=0)'],
                    stopIfTrue=True,
                    fill=PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
                    font=Font(color="999999")
                )
            )
            
            # Members at other thresholds (very light gray)
            ws.conditional_formatting.add(
                f'A{row_idx}:H{row_idx}',
                FormulaRule(
                    formula=[f'C{row_idx}<>Dashboard!$B$4'],
                    stopIfTrue=True,
                    fill=PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid"),
                    font=Font(color="CCCCCC")
                )
            )
        
        # Add count of active members
        ws['D3'] = "Active Members:"
        ws['D3'].font = Font(bold=True, size=11)
        ws['E3'] = f'=COUNTIFS(I6:I{last_row}, 1)'
        ws['E3'].font = Font(bold=True, size=11, color="4472C4")
        ws['E3'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws['E3'].border = border_thin
        ws['E3'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Hide the filter column
        ws.column_dimensions['I'].width = 5
        ws.column_dimensions['I'].hidden = True
        
    else:
        # Original filtering logic for non-evolution data
        ws['H5'] = "Filter"
        ws['H5'].fill = header_fill
        ws['H5'].font = header_font
        ws['H5'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add filter formula to all data rows
        for row_idx in range(6, last_row + 1):
            ws[f'H{row_idx}'] = f'=IF(A{row_idx}=Dashboard!$B$4, 1, 0)'
        
        # Add conditional formatting to highlight visible rows
        for row_idx in range(6, last_row + 1):
            # Highlight rows that match the threshold
            ws.conditional_formatting.add(
                f'A{row_idx}:G{row_idx}',
                FormulaRule(
                    formula=[f'=$H{row_idx}=1'],
                    stopIfTrue=True,
                    fill=PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                )
            )
            
            # Dim rows that don't match
            ws.conditional_formatting.add(
                f'A{row_idx}:G{row_idx}',
                FormulaRule(
                    formula=[f'=$H{row_idx}=0'],
                    stopIfTrue=True,
                    fill=PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
                    font=Font(color="999999")
                )
            )
        
        # Add count of visible groups
        ws['D3'] = "Groups at this Threshold:"
        ws['D3'].font = Font(bold=True, size=11)
        ws['E3'] = f'=COUNTIF(H6:H{last_row}, 1)'
        ws['E3'].font = Font(bold=True, size=11, color="4472C4")
        ws['E3'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws['E3'].border = border_thin
        ws['E3'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column H (filter) to be narrow and hide it
        ws.column_dimensions['H'].width = 5
        ws.column_dimensions['H'].hidden = True
    
    # Auto-size columns
    for column_cells in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column_cells:
            # Skip merged cells
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        if column_letter:
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A6'


def _write_dashboard_sheet(ws, summary_df, title_font, header_fill, header_font, border_thin):
    """Write the Dashboard sheet with dropdown selector and dynamic metrics."""
    # Title
    ws['A1'] = "Threshold Explorer Dashboard"
    ws['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws.merge_cells('A1:E1')
    
    # Instructions
    ws['A2'] = "Select a threshold below to see how grouping changes"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:E2')
    
    # Threshold selector
    ws['A4'] = "Select Threshold:"
    ws['A4'].font = Font(bold=True, size=12)
    ws['A4'].alignment = Alignment(horizontal='right', vertical='center')
    
    # Get threshold values from summary
    threshold_values = sorted(summary_df['Threshold'].unique())
    default_threshold = threshold_values[len(threshold_values) // 2]  # Middle value
    
    # Create dropdown using Data Validation
    ws['B4'] = default_threshold
    ws['B4'].font = Font(bold=True, size=12)
    ws['B4'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws['B4'].border = border_thin
    ws['B4'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data validation for dropdown
    threshold_list = ','.join(str(t) for t in threshold_values)
    dv = DataValidation(type="list", formula1=f'"{threshold_list}"', allow_blank=False)
    dv.add(ws['B4'])
    ws.add_data_validation(dv)
    
    # Summary metrics section
    ws['A6'] = "Summary Metrics"
    ws['A6'].font = Font(bold=True, size=14)
    ws.merge_cells('A6:E6')
    
    # Metric labels and formulas
    metrics = [
        ("Groups Found", "B", "Groups Found"),
        ("Products in Groups", "C", "Products in Groups"),
        ("Largest Group Size", "D", "Largest Group Size"),
        ("Avg Group Similarity", "E", "Avg Group Similarity"),
        ("Singletons", "F", "Singletons")
    ]
    
    row = 8
    for metric_name, summary_col, _ in metrics:
        # Label
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        # Value (VLOOKUP formula)
        col_idx = summary_df.columns.get_loc(metric_name) + 1
        ws[f'B{row}'] = f'=VLOOKUP(B4,Summary!$A$4:$F$100,{col_idx},FALSE)'
        ws[f'B{row}'].font = Font(size=12, bold=True)
        ws[f'B{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws[f'B{row}'].border = border_thin
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
    
    # Chart section
    ws['A14'] = "Groups vs Threshold"
    ws['A14'].font = Font(bold=True, size=14)
    
    # Create line chart
    chart = LineChart()
    chart.title = "How Groups Change with Threshold"
    chart.style = 10
    chart.y_axis.title = "Number of Groups"
    chart.x_axis.title = "Similarity Threshold (%)"
    chart.height = 10
    chart.width = 20
    
    # Data for chart from Summary sheet
    data = Reference(ws.parent['Summary'], min_col=2, min_row=3, max_row=3 + len(summary_df))
    cats = Reference(ws.parent['Summary'], min_col=1, min_row=4, max_row=3 + len(summary_df))
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "A16")
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15


def _write_heatmap_sheet(ws, similarity_matrix, product_names, header_fill, header_font):
    """Write a similarity heatmap with conditional formatting."""
    # Title
    ws['A1'] = "Similarity Heatmap"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')
    
    ws['A2'] = "Color scale: Red (low similarity) → Yellow → Green (high similarity)"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:C2')
    
    # Headers
    ws['A4'] = "Product"
    ws['A4'].fill = header_fill
    ws['A4'].font = header_font
    
    # Write product names as column headers
    for col_idx, name in enumerate(product_names, start=2):
        cell = ws.cell(row=4, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
    
    # Write similarity matrix
    for row_idx, name in enumerate(product_names, start=5):
        # Row header (product name)
        ws.cell(row=row_idx, column=1, value=name)
        
        # Similarity values
        for col_idx in range(len(product_names)):
            value = similarity_matrix[row_idx - 5][col_idx]
            cell = ws.cell(row=row_idx, column=col_idx + 2, value=value)
            cell.number_format = '0.00'
    
    # Apply conditional formatting (color scale)
    data_range = f'B5:{chr(65 + len(product_names))}{4 + len(product_names)}'
    color_scale = ColorScaleRule(
        start_type='num', start_value=0, start_color='F8696B',
        mid_type='num', mid_value=50, mid_color='FFEB84',
        end_type='num', end_value=100, end_color='63BE7B'
    )
    ws.conditional_formatting.add(data_range, color_scale)
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    for col_idx in range(len(product_names)):
        ws.column_dimensions[chr(66 + col_idx)].width = 4
    
    # Freeze panes
    ws.freeze_panes = 'B5'
