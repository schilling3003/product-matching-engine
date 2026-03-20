"""
Create an Excel file with VBA macro code in a worksheet for easy copy-paste
This creates a ready-to-use Excel file where users can copy the macro code directly
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

def create_excel_with_macro_code():
    """Create an Excel file containing the VBA macro code in a worksheet"""
    
    # Read the VBA macro code
    macro_file = Path(__file__).parent / "Excel_Threshold_Explorer_Macro.bas"
    
    if not macro_file.exists():
        print(f"Error: Macro file not found at {macro_file}")
        return
    
    with open(macro_file, 'r', encoding='utf-8') as f:
        vba_code = f.read()
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "VBA Macro Code"
    
    # Set up styles
    title_font = Font(size=16, bold=True, color="FFFFFF")
    header_font = Font(size=12, bold=True)
    code_font = Font(name="Consolas", size=10)
    
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    code_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws['A1'] = "Excel Threshold Explorer - VBA Macro Setup"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws.merge_cells('A1:D1')
    
    # Add instructions
    instructions = [
        ("How to Install and Use the Macro:", header_font, header_fill),
        ("", None, None),
        ("1. Open this Excel file", None, None),
        ("2. Press Alt+F11 to open the VBA Editor", None, None),
        ("3. In the VBA Editor, click Insert > Module", None, None),
        ("4. Copy all the code from column B below", None, None),
        ("5. Paste the code into the new module", None, None),
        ("6. Close the VBA Editor", None, None),
        ("7. Press Alt+F8, select 'SetupThresholdExplorer', and click Run", None, None),
        ("8. Save the file as Excel Macro-Enabled Workbook (.xlsm)", None, None),
        ("", None, None),
        ("Important Notes:", header_font, header_fill),
        ("• You must have Power Pivot enabled (Excel 2016+)", None, None),
        ("• Enable macros when prompted", None, None),
        ("• Import your data from the enhanced export first", None, None),
        ("", None, None),
        ("VBA Code (copy everything below this line):", header_font, header_fill)
    ]
    
    row = 3
    for text, font, fill in instructions:
        if text:
            ws[f'A{row}'] = text
            ws[f'A{row}'].font = font or Font(size=11)
            ws[f'A{row}'].fill = fill or PatternFill()
            ws[f'A{row}'].alignment = Alignment(vertical='top')
            ws[f'A{row}'].border = thin_border
            ws.merge_cells(f'A{row}:D{row}')
        row += 1
    
    # Add the VBA code
    ws['B' + str(row)] = "VBA CODE START"
    ws['B' + str(row)].font = header_font
    row += 1
    
    # Split VBA code into lines and add to worksheet
    lines = vba_code.split('\n')
    
    for line in lines:
        # Handle very long lines by splitting them
        if len(line) > 32767:  # Excel cell limit
            # Split long lines
            chunks = [line[i:i+3000] for i in range(0, len(line), 3000)]
            for chunk in chunks:
                ws[f'B{row}'] = chunk
                ws[f'B{row}'].font = code_font
                ws[f'B{row}'].fill = code_fill
                ws[f'B{row}'].alignment = Alignment(vertical='top', wrap_text=True)
                ws[f'B{row}'].border = thin_border
                row += 1
        else:
            ws[f'B{row}'] = line
            ws[f'B{row}'].font = code_font
            ws[f'B{row}'].fill = code_fill
            ws[f'B{row}'].alignment = Alignment(vertical='top', wrap_text=True)
            ws[f'B{row}'].border = thin_border
            row += 1
    
    # Add a "Select All" button area
    row += 2
    ws[f'A{row}'] = "Quick Select All Code:"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].border = thin_border
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = "1. Click on cell B" + str(row - len(lines) - 1)
    ws[f'A{row}'].font = Font(size=11)
    row += 1
    
    ws[f'A{row}'] = "2. Scroll down to the last line of code"
    ws[f'A{row}'].font = Font(size=11)
    row += 1
    
    ws[f'A{row}'] = "3. Hold Shift and click the last cell"
    ws[f'A{row}'].font = Font(size=11)
    row += 1
    
    ws[f'A{row}'] = "4. Press Ctrl+C to copy all"
    ws[f'A{row}'].font = Font(size=11)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    
    # Freeze the top rows
    ws.freeze_panes = 'A20'
    
    # Add a second sheet with quick reference
    ws2 = wb.create_sheet("Quick Reference")
    
    # Quick reference content
    quick_ref = [
        ["Macro Functions", "Description"],
        ["SetupThresholdExplorer", "Complete automated setup of the threshold explorer"],
        ["UpdateAnalysis", "Update analysis with new threshold value"],
        ["ExportCurrentThreshold", "Export analysis for current threshold"],
        ["ResetWorkbook", "Remove all added sheets and reset to original"],
        ["", ""],
        ["Keyboard Shortcuts", ""],
        ["Alt+F11", "Open VBA Editor"],
        ["Alt+F8", "Open Macro Dialog"],
        ["F5", "Run selected macro in VBA"],
        ["Alt+P+C", "Open Power Pivot (if available)"],
        ["", ""],
        ["Troubleshooting", ""],
        ["Macro not found", "Make sure you've imported the code into a module"],
        ["Power Pivot error", "Check if Power Pivot is enabled in Excel options"],
        ["Runtime error", "Ensure all required sheets exist in your workbook"]
    ]
    
    for r, (key, value) in enumerate(quick_ref, 1):
        ws2[f'A{r}'] = key
        ws2[f'B{r}'] = value
        ws2[f'A{r}'].font = Font(bold=True) if r == 1 or r == 7 or r == 11 or r == 14 else Font()
        ws2[f'B{r}'].font = Font(bold=True) if r == 1 else Font()
    
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 60
    
    # Save the file
    output_file = Path(__file__).parent / "Threshold_Explorer_Ready_to_Use.xlsx"
    wb.save(output_file)
    
    print(f"Created Excel file: {output_file}")
    print("\nThis file contains:")
    print("- Step-by-step instructions")
    print("- Complete VBA macro code ready to copy")
    print("- Quick reference sheet")
    print("\nUsers can simply copy the code from column B and paste into VBA editor!")

if __name__ == "__main__":
    create_excel_with_macro_code()
